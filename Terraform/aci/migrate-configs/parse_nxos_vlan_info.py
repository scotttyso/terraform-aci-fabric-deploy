#!/usr/bin/env python3

import ipaddress
import openpyxl
import pandas as pd
import numpy as np
import os, re, sys, traceback, validators
from openpyxl import Workbook
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side 
from openpyxl.utils.dataframe import dataframe_to_rows

# Define Regular Expressions to be used in function definations and searches
re_dhcp = re.compile(r'^  ip dhcp relay address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}) $\n')
re_desc = re.compile('^  description (.+)$\n')
re_host = re.compile('^hostname (.+)$\n')
re_hsv4 = re.compile(r'^    ip (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})$\n')
re_hsv4s = re.compile(r'^    ip (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}) secondary$\n')
re_ipv4 = re.compile(r'^  ip address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}(?:/\d{1,2}|))$\n')
re_ipv4s = re.compile(r'^  ip address (\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3}(?:/\d{1,2}|)) secondary$\n')
re_ivln = re.compile(r'^interface Vlan(\d+)$\n')
re_mtu_ = re.compile(r'^  mtu (\d+)$\n')
re_vlan = re.compile(r'^vlan (\d{1,4})$\n')
re_vlnm = re.compile('^  name (.+)$\n')
re_vlst = re.compile(r'^vlan (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_vrf_ = re.compile('^  vrf member (.+)$\n')

def function_wr_vlan(vlan):
    if vlan < 10:
        vlan = str(vlan)
        wr_vlan.write('v000{}_bd\n'.format(vlan))
    elif vlan < 100:
        vlan = str(vlan)
        wr_vlan.write('v00{}_bd\n'.format(vlan))
    elif vlan < 1000:
        vlan = str(vlan)
        wr_vlan.write('v0{}_bd\n'.format(vlan))
    else:
        vlan = str(vlan)
        wr_vlan.write('v{}_bd\n'.format(vlan))

def function_wr_name(vlan,name):
    if vlan < 10:
        wr_name.write('v000{}_bd\t{}\n'.format(vlan, name))
    elif vlan < 100:
        vlan = str(vlan)
        wr_name.write('v00{}_bd\t{}\n'.format(vlan, name))
    elif vlan < 1000:
        vlan = str(vlan)
        wr_name.write('v0{}_bd\t{}\n'.format(vlan, name))
    else:
        vlan = str(vlan)
        wr_name.write('v{}_bd\t{}\n'.format(vlan, name))

def function_vlan_to_bd(vlan):
    if vlan < 10:
        vlan = str(vlan)
        bd = 'v000' + vlan + '_bd'
        return bd
    elif vlan < 100:
        vlan = str(vlan)
        bd = 'v00' + vlan + '_bd'
        return bd
    elif vlan < 1000:
        vlan = str(vlan)
        bd = 'v0' + vlan + '_bd'
        return bd
    else:
        vlan = str(vlan)
        bd = 'v' + vlan + '_bd'
        return bd

def function_expand_vlst(vlst):
    vlist = str_vlst.split(',')
    for v in vlist:
        if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
            a,b = v.split('-')
            a = int(a)
            b = int(b)
            vrange = range(a,b+1)
            for vl in vrange:
                function_wr_vlan(vl)
        elif re.fullmatch('^\\d{1,4}$', v):
            v = int(v)
            function_wr_vlan(v)
    

# Start by Creating Empty Variables
hostname = ''
str_dhcp = ''
str_desc = ''
str_hsv4 = ''
str_hsv4s = ''
str_ipv4 = ''
str_ipv4s = ''
str_ivln = ''
str_mtu_ = '9000'
str_vlan = ''
str_vlst = ''
str_vlnm = ''
str_vrf_ = 'default'

# Import the Configuration File
config_file = sys.argv[1]
try:
    if os.path.isfile(config_file):
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   {config_file} exists.  Beginning Script Execution...')
        print(f'\n-----------------------------------------------------------------------------\n')
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   {config_file} does not exist.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
except IOError:
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   {config_file} does not exist.  Exiting....')
    print(f'\n-----------------------------------------------------------------------------\n')
    exit()

file = open(config_file, 'r')
wr_vlan = open('vlan_list.csv', 'w')
wr_name = open('vlan_name.csv', 'w')
wr_dhcp = open('dhcp.csv', 'w')


bd1 = Side(style="thick", color="8EA9DB")
bd2 = Side(style="medium", color="8EA9DB")
wsh1 = NamedStyle(name="wsh1")
wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
wsh1.font = Font(bold=True, size=15, color="FFFFFF")
wsh2 = NamedStyle(name="wsh2")
wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
wsh2.fill = PatternFill("solid", fgColor="305496")
wsh2.font = Font(bold=True, size=15, color="FFFFFF")
ws_odd = NamedStyle(name="ws_odd")
ws_odd.alignment = Alignment(horizontal="center", vertical="center")
ws_odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_odd.fill = PatternFill("solid", fgColor="D9E1F2")
ws_odd.font = Font(bold=False, size=12, color="44546A")
ws_even = NamedStyle(name="ws_even")
ws_even.alignment = Alignment(horizontal="center", vertical="center")
ws_even.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_even.font = Font(bold=False, size=12, color="44546A")


wb = Workbook()
wb.add_named_style(wsh1)
wb.add_named_style(wsh2)
wb.add_named_style(ws_odd)
wb.add_named_style(ws_even)

dest_file = 'export_vlan_info.xlsx'
ws1 = wb.active
ws1.title = "Tenant"
ws2 = wb.create_sheet(title = "VRF")
ws3 = wb.create_sheet(title = "Bridge_Domain")
ws4 = wb.create_sheet(title = "Gateway")
ws5 = wb.create_sheet(title = "DHCP Relay")
ws2 = wb["VRF"]
ws3 = wb["Bridge_Domain"]
ws4 = wb["Gateway"]
ws5 = wb["DHCP Relay"]
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 40
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 40
ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 40
ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 25
ws4.column_dimensions['H'].width = 20
ws4.column_dimensions['I'].width = 40
ws5.column_dimensions['A'].width = 15
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 20
ws5.column_dimensions['E'].width = 40

data = ['Type','Tenant Name','Description']
ws1.append(data)
for cell in ws1["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','Description']
ws2.append(data)
for cell in ws2["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','Bridge Domain','Extend Outside ACI','Description']
ws3.append(data)
for cell in ws3["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','Bridge Domain','Gateway IPv4','Gateway_Type','Advertise External', 'Layer3 Out','Description']
ws4.append(data)
for cell in ws4["1:1"]:
    cell.style = 'wsh2'
data = ['Type','Tenant','VRF Name','IPv4 Address','Description']
ws5.append(data)
for cell in ws5["1:1"]:
    cell.style = 'wsh2'
ws1_row_count = 2
ws2_row_count = 2
ws3_row_count = 2
ws4_row_count = 2
ws5_row_count = 2

# Read the Conifguration File and Gather Vlan Information
lines = file.readlines()

line_count = 0

for line in lines:
    if re.fullmatch(re_vlst, line):
        # Matched the VLAN List... Now Parse for Data Export
        str_vlst = re.fullmatch(re_vlst, line).group(1)
        # Expand VLAN Ranges into Full VLAN List
        function_expand_vlst(str_vlst)
        line_count += 1
    elif re.fullmatch(re_host, line):
        hostname = re.fullmatch(re_host, line).group(1)
    elif re.fullmatch(re_vlan, line):
        # Matched a VLAN... Now Parse for Data Export
        str_vlan = int(re.fullmatch(re_vlan, line).group(1))
        line_count += 1
    elif re.fullmatch(re_vlnm, line):
        # Matched VLAN Name... Now Parse for Data Export
        str_vlnm = re.fullmatch(re_vlnm, line).group(1)
        function_wr_name(str_vlan,str_vlnm)
        line_count += 1
    elif re.fullmatch(re_ivln, line):
        # Matched an Interface VLAN... Now Parse for Data Export
        str_ivln = int(re.fullmatch(re_ivln, line).group(1))
        line_count += 1
    elif re.fullmatch(re_mtu_, line):
        # Matched the Interface MTU... Now Parse for Data Export
        str_mtu_ = re.fullmatch(re_mtu_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_vrf_, line):
        # Matched a VRF Context... Now Parse for Data Export
        str_vrf_ = re.fullmatch(re_vrf_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_ipv4, line):
        # Matched an IPv4 Address/prefix... Now Parse for Data Export
        str_ipv4 = re.fullmatch(re_ipv4, line).group(1)
        line_count += 1
    elif re.fullmatch(re_ipv4s, line):
        # Matched an IPv4 Secondary Address/prefix... Now Parse for Data Export
        str_ipv4s = re.fullmatch(re_ipv4s, line).group(1)
        line_count += 1
    elif re.fullmatch(re_hsv4, line):
        # Matched an HSRP IPv4 Address... Now Parse for Data Export
        str_hsv4 = re.fullmatch(re_hsv4, line).group(1)
        line_count += 1
    elif re.fullmatch(re_hsv4s, line):
        # Matched an HSRP IPv4 Secondary Address/prefix... Now Parse for Data Export
        str_hsv4s = re.fullmatch(re_hsv4s, line).group(1)
        line_count += 1
    elif re.fullmatch(re_dhcp, line):
        # Matched an IPv4 DHCP Relay definition... Now Parse for Data Export
        str_dhcp = re.fullmatch(re_dhcp, line).group(1)
        wr_dhcp.write('{},{}\n'.format(str_vrf_, str_dhcp))
        line_count += 1
    elif re.fullmatch(re_desc, line):
        # Found a Description on the Interface
        str_desc = re.fullmatch(re_desc, line).group(1)
        line_count += 1
    elif line == "\n":
        # Found blank line, which means the end of the interface, time to create the output
        if not str_ipv4:
            line_count += 1
        elif not str_ivln:
            line_count += 1
        elif str_ipv4:
            bd = function_vlan_to_bd(str_ivln)
            if str_hsv4:
                a,b = str_ipv4.split('/')
                gtwy = str(str_hsv4) + '/' + str(b)
            else:
                gtwy = str(str_ipv4)
            data = ['subnet_add','',str_vrf_,bd,gtwy,'primary','','',str_desc]
            ws4.append(data)
            rc = '{}:{}'.format(ws4_row_count, ws4_row_count)
            for cell in ws4[rc]:
                if ws4_row_count % 2 == 0:
                    cell.style = 'ws_even'
                else:
                    cell.style = 'ws_odd'
            ws4_row_count += 1
            if str_ipv4s:
                if str_hsv4s:
                    a,b = str_ipv4s.split('/')
                    gtwy = str(str_hsv4s) + '/' + str(b)
                else:
                    gtwy = str(str_ipv4)
                data = ['subnet_add','',str_vrf_,bd,gtwy,'secondary','','',str_desc]
                ws4.append(data)
                rc = '{}:{}'.format(ws4_row_count, ws4_row_count)
                for cell in ws4[rc]:
                    if ws4_row_count % 2 == 0:
                        cell.style = 'ws_even'
                    else:
                        cell.style = 'ws_odd'
                ws4_row_count += 1
            line_count += 1
        
        # Reset the Variables back to Blank
        str_dhcp = ''
        str_desc = ''
        str_hsv4 = ''
        str_hsv4s = ''
        str_ipv4 = ''
        str_ipv4s = ''
        str_ivln = ''
        str_mtu_ = '9000'
        str_vlan = ''
        str_vlst = ''
        str_vlnm = ''
        str_vrf_ = 'default'
    else:
        line_count += 1


file.close()
wr_vlan.close()
wr_name.close()
wr_dhcp.close()

#Get VLAN's that don't have a name and those that do and combine into one file
bg_list1 = open('vlan_list.csv', 'r') 
bg_list2 = open('vlan_name.csv', 'r')
bg_list3 = open('vlan_comb.csv', 'w')
vlan_lines = bg_list1.readlines()
name_lines = bg_list2.readlines()
for lineg1 in vlan_lines:
    matched = 0
    lineg1 = lineg1.strip()
    for lineg2 in name_lines:
        lineg2 = lineg2.strip()
        if lineg1 in lineg2:
            matched +=1
    if matched == 0:
        bg_list3.write('{}\n'.format(lineg1))
for line in name_lines:
    line.strip()
    bg_list3.write('{}'.format(line))

bg_list1.close()
bg_list2.close()
bg_list3.close()

#Sort the combined VLANs in final output file
bg_list3 = open('vlan_comb.csv', 'r')
bddm = bg_list3.readlines()
bddm.sort()
for line in range(len(bddm)):
    bddm[line]
    if re.search('\t', bddm[line]):
        bd,descr = bddm[line].split('\t')
    else:
        bd = bddm[line]
        descr = ''
    data = ['bd_add','','',bd,'yes',descr]
    ws3.append(data)
    rc = '{}:{}'.format(ws3_row_count, ws3_row_count)
    for cell in ws3[rc]:
        if ws3_row_count % 2 == 0:
            cell.style = 'ws_even'
        else:
            cell.style = 'ws_odd'
    ws3_row_count += 1
    #bg_list4.write('bd_add,extend_out,{}'.format(bddm[line]))

bg_list3.close()

dhcp_relay_uniq = 'cat dhcp.csv | sort | uniq > dhcp_sort.csv'
os.system(dhcp_relay_uniq)

file_relays = open('dhcp_sort.csv', 'r')
read_relays = file_relays.readlines()
for line in read_relays:
    vrf,relay_ip = line.split(',')
    data = ['dhcp_relay','Tenant',vrf,relay_ip]
    ws5.append(data)
    rc = '{}:{}'.format(ws5_row_count, ws5_row_count)
    for cell in ws5[rc]:
        if ws5_row_count % 2 == 0:
            cell.style = 'ws_even'
        else:
            cell.style = 'ws_odd'
    ws5_row_count += 1
file_relays.close()
remove_extra_files = 'rm dhcp.csv dhcp_sort.csv vlan_comb.csv vlan_list.csv vlan_name.csv'
os.system(remove_extra_files)

# Save the Excel Workbook
wb.save(dest_file)

if not hostname == '':
    rename_excel = 'mv export_vlan_info.xlsx {}_export_vlan_inf.xlsx'.format(hostname)
    os.system(rename_excel)

#End Script
print(f'\n-----------------------------------------------------------------------------\n')
print(f'   Completed Running Script.  Exiting....')
print(f'\n-----------------------------------------------------------------------------\n')
exit()