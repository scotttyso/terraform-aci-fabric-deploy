#!/usr/bin/env python3

import ipaddress
import openpyxl
import pandas as pd
import numpy as np
import os, re, sys, traceback, validators
from openpyxl import Workbook
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side 
from openpyxl.utils.dataframe import dataframe_to_rows

re_bpdu = re.compile('^  spanning-tree bpduguard enable$\n')
re_cdpe = re.compile('^  cdp enable$\n')
re_desc = re.compile('^  description (.+)$\n')
re_host = re.compile('^hostname (.+)$\n')
re_intf = re.compile(r'^interface ((port\-channel\d+|Ethernet\d+[\d\/]+))$\n')
re_ldpr = re.compile('^  lldp transmit$\n')
re_ldpt = re.compile('^  lldp receive$\n')
re_mtu_ = re.compile(r'^  mtu (\d+)$\n')
re_poch = re.compile(r'^  channel-group (\d+) mode ((active|on|passive))$\n')
re_swav = re.compile(r'^  switchport access vlan (\d+)$\n')
re_swma = re.compile('^  switchport mode access$\n')
re_swmt = re.compile('^  switchport mode trunk$\n')
re_swpt = re.compile('^  switchport$\n')
re_tknv = re.compile(r'^  switchport trunk native vlan (\d{1,4})$\n')
re_tkv1 = re.compile(r'^  switchport trunk allowed vlan (\d{1,4}[\-,]+.+\d{1,4})$\n')
re_tkv2 = re.compile(r'^  switchport trunk allowed vlan (\d{1,4})$\n')
re_vpc_ = re.compile(r'^  vpc ((\d+|peer\-link))$\n')

# Import the Configuration File
config_file = sys.argv[1]
try:
    if os.path.isfile(config_file):
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   {config_file} exists.  Beginning Script Execution...')
        print(f'\n-----------------------------------------------------------------------------\n')
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   {config_file} does not exist.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
except IOError:
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   {config_file} does not exist.  Exiting....')
    print(f'\n-----------------------------------------------------------------------------\n')
    exit()

file = open(config_file, 'r')
wr_poch = open('int_poch.csv', 'w')

def func_wr_poch(str_host, str_intf, str_vpc_, str_mtu_, str_swmd, str_swav, str_tknv, str_tkvl, str_desc):
    wr_poch.write('{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n'.format(str_host, str_intf, str_vpc_, str_mtu_, str_swmd, str_swav, str_tknv, str_tkvl, str_desc))


# Start by Creating Default Variable Values
str_bpdg = 'no'
str_cdp_ = 'no'
str_desc = ''
str_host = 'undefined'
str_intf = ''
str_lldr = 'no'
str_lldt = 'no'
str_mtu_ = ''
str_poch = 'n/a'
str_pomd = 'n/a'
str_swav = 'n/a'
str_swmd = 'access'
str_swpt = 'no'
str_tknv = 'n/a'
str_tkvl = 'n/a'
str_vpc_ = 'n/a'

bd1 = Side(style="thick", color="8EA9DB")
bd2 = Side(style="medium", color="8EA9DB")
wsh1 = NamedStyle(name="wsh1")
wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
wsh1.font = Font(bold=True, size=15, color="FFFFFF")
wsh2 = NamedStyle(name="wsh2")
wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
wsh2.fill = PatternFill("solid", fgColor="305496")
wsh2.font = Font(bold=True, size=15, color="FFFFFF")
ws_odd = NamedStyle(name="ws_odd")
ws_odd.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
ws_odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_odd.fill = PatternFill("solid", fgColor="D9E1F2")
ws_odd.font = Font(bold=False, size=12, color="44546A")
ws_even = NamedStyle(name="ws_even")
ws_even.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
ws_even.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
ws_even.font = Font(bold=False, size=12, color="44546A")

wb = Workbook()
wb.add_named_style(wsh1)
wb.add_named_style(wsh2)
wb.add_named_style(ws_odd)
wb.add_named_style(ws_even)

dest_file = 'migrate_interfaces.xlsx'
ws1 = wb.active
ws1.title = "Migrate Interfaces"
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 10
ws1.column_dimensions['I'].width = 10
ws1.column_dimensions['J'].width = 15
ws1.column_dimensions['K'].width = 17
ws1.column_dimensions['L'].width = 40
ws1.column_dimensions['M'].width = 12
ws1.column_dimensions['N'].width = 12
ws1.column_dimensions['O'].width = 12
ws1.column_dimensions['P'].width = 12
ws1.column_dimensions['Q'].width = 40
ws1.column_dimensions['R'].width = 40

data = ['Type','New Host','New Interface','Current Host','Current Interface','Port Type','port-channel ID','VPC Id','MTU','Switchport Mode',\
    'Access or Native VLAN','Trunk Allowed VLANs','CDP Enabled','LLDP Receive','LLDP Transmit','BPDU Guard','Port-Channel Description',\
    'Port Description']
ws1.append(data)
for cell in ws1["1:1"]:
    cell.style = 'wsh2'
ws1_row_count = 2

# Read the Conifguration File and Gather Vlan Information
lines = file.readlines()

line_count = 0
ethn_count = 0
for line in lines:
    if re.fullmatch(re_host, line):
        str_host = re.fullmatch(re_host, line).group(1)
        line_count += 1
    elif re.fullmatch(re_intf, line):
        str_intf = re.fullmatch(re_intf, line).group(1)
        line_count += 1
    elif re.fullmatch(re_bpdu, line):
        str_bpdg = 'yes'
        line_count += 1
    elif re.fullmatch(re_cdpe, line):
        str_cdp_ = 'yes'
        line_count += 1
    elif re.fullmatch(re_ldpr, line):
        str_lldr = 'yes'
        line_count += 1
    elif re.fullmatch(re_ldpt, line):
        str_lldt = 'yes'
        line_count += 1
    elif re.fullmatch(re_swav, line):
        str_swav = re.fullmatch(re_swav, line).group(1)
    elif re.fullmatch(re_swma, line):
        str_swmd = 'access'
    elif re.fullmatch(re_swmt, line):
        str_swmd = 'trunk'
    elif re.fullmatch(re_tknv, line):
        str_tknv = re.fullmatch(re_tknv, line).group(1)
        line_count += 1
    elif re.fullmatch(re_tkv1, line):
        str_tkvl = re.fullmatch(re_tkv1, line).group(1)
        line_count += 1
    elif re.fullmatch(re_tkv2, line):
        str_tkvl = re.fullmatch(re_tkv2, line).group(1)
        line_count += 1
    elif re.fullmatch(re_swpt, line):
        str_swpt = 'yes'
    elif re.fullmatch(re_mtu_, line):
        str_mtu_ = re.fullmatch(re_mtu_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_poch, line):
        str_poch = re.fullmatch(re_poch, line).group(1)
        str_pomd = re.fullmatch(re_poch, line).group(2)
        line_count += 1
    elif re.fullmatch(re_vpc_, line):
        str_vpc_ = re.fullmatch(re_vpc_, line).group(1)
        line_count += 1
    elif re.fullmatch(re_desc, line):
        str_desc = re.fullmatch(re_desc, line).group(1)
        line_count += 1
    elif line == "\n":
        # Found blank line, which means the end of the interface, time to create the output
        if 'channel' in str_intf:
            if str_swpt == 'yes':
                func_wr_poch(str_host, str_intf, str_vpc_, str_mtu_, str_swmd, str_swav, str_tknv, str_tkvl, str_desc)
        elif 'Ethernet' in str_intf:
            if ethn_count == 0:
                wr_poch.close()
                read_poch = open('int_poch.csv', 'r')
                po_lines = read_poch.readlines()
                ethn_count += 1
            if str_swpt == 'yes':
                if re.search(r'(\d+|peer)', str_poch):
                    for line in po_lines:
                        x = line.split('\t')
                        desc = x[8].strip()
                        if str_poch in x[1]:
                            if str_swmd == 'access':
                                swav = x[5]
                            else:
                                swav = x[6]
                            if x[2] == 'n/a':
                                po_type = 'pc'
                            else:
                                po_type = 'vpc'
                            data = ['intf_add','','',str_host,str_intf,po_type,str_poch,x[2],x[3],x[4],swav,x[7],str_cdp_,str_lldr,
                                    str_lldt,str_bpdg,desc,str_desc]
                            ws1.append(data)
                            rc = '{}:{}'.format(ws1_row_count, ws1_row_count)
                            for cell in ws1[rc]:
                                if ws1_row_count % 2 == 0:
                                    cell.style = 'ws_even'
                                else:
                                    cell.style = 'ws_odd'
                            ws1_row_count += 1
                else:
                    po_type = 'ap'
                    if str_swmd == 'access':
                        swav = str_swav
                    else:
                        swav = str_tknv
                    data = ['intf_add','','',str_host,str_intf,po_type,str_poch,str_vpc_,str_mtu_,str_swmd,swav,str_tkvl,str_cdp_,str_lldr,
                            str_lldt,str_bpdg,'n/a',str_desc]
                    ws1.append(data)
                    rc = '{}:{}'.format(ws1_row_count, ws1_row_count)
                    for cell in ws1[rc]:
                        if ws1_row_count % 2 == 0:
                            cell.style = 'ws_even'
                        else:
                            cell.style = 'ws_odd'
                    ws1_row_count += 1

        # Reset the Variables back to Default
        str_bpdg = 'no'
        str_cdp_ = 'no'
        str_desc = ''
        str_intf = ''
        str_lldr = 'no'
        str_lldt = 'no'
        str_mtu_ = 'n/a'
        str_poch = 'n/a'
        str_pomd = 'n/a'
        str_swav = '1'
        str_swmd = 'access'
        str_swpt = 'no'
        str_tknv = '1'
        str_tkvl = 'n/a'
        str_vpc_ = 'n/a'
        line_count += 1
    else:
        line_count += 1


file.close()

remove_poch_file = 'rm int_poch.csv'
os.system(remove_poch_file)

# Save the Excel Workbook
wb.save(dest_file)

if not str_host == '':
    rename_excel = 'mv migrate_interfaces.xlsx {}_migrate_interfaces.xlsx'.format(str_host)
    os.system(rename_excel)

#End Script
print(f'\n-----------------------------------------------------------------------------\n')
print(f'   Completed Running Script.  Exiting....')
print(f'\n-----------------------------------------------------------------------------\n')
exit()